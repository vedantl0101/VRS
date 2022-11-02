from lib2to3.pygram import python_symbols
from tkinter import *
from tkinter import ttk
import random
import time
import datetime
from PIL import ImageTk, Image
from tkinter import messagebox as ms
import sqlite3
from tkinter.font import BOLD
import pandas as pd
import openpyxl as op
wb=op.load_workbook("Vehicle.xlsx")

Item4 = 0


# make database and users (if not exists already) table at programme start up
with sqlite3.connect('Users.db') as db:
    c = db.cursor()

c.execute('CREATE TABLE IF NOT EXISTS user (username TEXT NOT NULL ,password TEXT NOT NULL)')
db.commit()
db.close()

#main Class
class user:
    def __init__(self,master):
    	# Window 
        self.master = master
        # Some Usefull variables
        self.username = StringVar()
        self.password = StringVar()
        self.n_username = StringVar()
        self.n_password = StringVar()
        #Create Widgets
        self.widgets()

    #Login Function
    def login(self):
    	#Establish Connection
        with sqlite3.connect('Users.db') as db:
            c = db.cursor()

        #Find user If there is any take proper action
        find_user = ('SELECT * FROM user WHERE username = ? and password = ?')
        c.execute(find_user,[(self.username.get()),(self.password.get())])
        result = c.fetchall()
        if result:
            self.logf.pack_forget()
            self.head['text'] = "Welcome " + self.username.get()
            self.head.configure(fg="black")
            self.head.pack(fill=X)
            application = List(root)
            
        else:
            ms.showerror('Oops!','Username Not Found.')
            
    def new_user(self):
    	#Establish Connection
        with sqlite3.connect('Users.db') as db:
            c = db.cursor()

        #Find Existing username if any take proper action
        find_user = ('SELECT * FROM user WHERE username = ?')
        c.execute(find_user,[(self.username.get())])        
        if c.fetchall():
            ms.showerror('Error!','Username Already Taken!')
        else:
            ms.showinfo('Success!','Account Created!')
            self.log()
        #Create New Account 
        insert = 'INSERT INTO user(username,password) VALUES(?,?)'
        c.execute(insert,[(self.n_username.get()),(self.n_password.get())])
        db.commit()

        #Frame Packing Methords
    def log(self):
        self.username.set('')
        self.password.set('')
        self.crf.pack_forget()
        self.head['text'] = 'Login'
        self.logf.pack()
    def cr(self):
        self.n_username.set('')
        self.n_password.set('')
        self.logf.pack_forget()
        self.head['text'] = 'Create Account'
        self.crf.pack()
        
    #Draw Widgets
    def widgets(self):
        self.head = Label(self.master,text = 'System',font = ('',30, BOLD),pady = 10)
        self.head.pack()
        self.logf = Frame(self.master,padx =10,pady = 10)
        Label(self.logf,text = 'Username: ',font = ('',20),pady=5,padx=5).grid(sticky = W)
        Entry(self.logf,textvariable = self.username,bd = 5,font = ('',15)).grid(row=0,column=1)
        Label(self.logf,text = 'Password: ',font = ('',20),pady=5,padx=5).grid(sticky = W)
        Entry(self.logf,textvariable = self.password,bd = 5,font = ('',15),show = '*').grid(row=1,column=1)
        Button(self.logf,text = ' Login ',bd = 3 ,font = ('',15),padx=5,pady=5,command=self.login).grid()
        Button(self.logf,text = ' Create Account ',bd = 3 ,font = ('',15),padx=5,pady=5,command=self.cr).grid(row=2,column=1)
        self.logf.pack()
        
        self.crf = Frame(self.master,padx =10,pady = 10)
        Label(self.crf,text = 'Username: ',font = ('',20),pady=5,padx=5).grid(sticky = W)
        Entry(self.crf,textvariable = self.n_username,bd = 5,font = ('',15)).grid(row=0,column=1)
        Label(self.crf,text = 'Password: ',font = ('',20),pady=5,padx=5).grid(sticky = W)
        Entry(self.crf,textvariable = self.n_password,bd = 5,font = ('',15),show = '*').grid(row=1,column=1)
        Button(self.crf,text = 'Create Account',bd = 3 ,font = ('',15),padx=5,pady=5,command=self.new_user).grid()
        Button(self.crf,text = 'Go to Login',bd = 3 ,font = ('',15),padx=5,pady=5,command=self.log).grid(row=2,column=1)

class List:
    def __init__(self,master):
        self.master = master
        self.number=StringVar()
        self.widget()
    def widget(self):
        self.head = Label(self.master,text = 'Vehicle Rental System',font = ('',30),pady = 10)
        self.head.pack()
        self.logf = Frame(self.master,padx =10,pady = 10)
        Button(self.logf,text = ' List of Vehicle ',bd = 3 ,font = ('',15),padx=5,pady=5,command=self.LOV).grid()
        Button(self.logf,text = ' Return A Vehicle ',bd = 3 ,font = ('',15),padx=5,pady=5,command=self.RAV).grid()
        Button(self.logf,text = ' Logout ',bd = 3 ,font = ('',15),padx=5,pady=5,command=self.logout).grid()
        
        self.logf.pack()
    def LOV(self):
        self.head.pack_forget()
        self.logf.pack_forget()
        self.crof = Frame(self.master,padx =20,pady = 10)
        Label(self.crof,text = 'Rent The Vehicle',font = ('',20),pady=5,padx=5).grid(sticky = W)
    
        Button(self.crof,text = ' Bike ',bd = 3 ,font = ('',15),padx=5,pady=5,command=self.Bike).grid()
        Button(self.crof,text = ' Car ',bd = 3 ,font = ('',15),padx=5,pady=5,command=self.Car).grid()
        Button(self.crof,text = ' Cycle ',bd = 3 ,font = ('',15),padx=5,pady=5,command=self.Cycle).grid()
        Button(self.crof,text = ' Scooter ',bd = 3 ,font = ('',15),padx=5,pady=5,command=self.Scooter).grid()
        
        self.crof.pack()
    
    def Bike(self):
        self.head.pack_forget()
        sh1=wb["Bike"]
        self.crof.pack_forget()
        self.head = Label(self.master,text = 'Choose A Bike',font = ('',30),pady = 10)
        self.head.pack()
        self.logf = Frame(self.master,padx =10,pady = 10)
        Radiobutton(self.logf,text = sh1.cell(2,2).value,bd = 3 ,font = ('',15),padx=5,pady=5,command=lambda:self.trv( sh1.cell(2,2).value)).grid(sticky=W)
        Radiobutton(self.logf,text = sh1.cell(3,2).value,bd = 3 ,font = ('',15),padx=5,pady=5,command=lambda :self.trv( sh1.cell(3,2).value)).grid(sticky=W)
        Radiobutton(self.logf,text = sh1.cell(4,2).value,bd = 3 ,font = ('',15),padx=5,pady=5,command=lambda: self.trv( sh1.cell(4,2).value)).grid(sticky=W)
        Radiobutton(self.logf,text = sh1.cell(5,2).value,bd = 3 ,font = ('',15),padx=5,pady=5,command=lambda: self.trv( sh1.cell(5,2).value)).grid(sticky=W)
        Radiobutton(self.logf,text = sh1.cell(6,2).value,bd = 3 ,font = ('',15),padx=5,pady=5,command=lambda: self.trv( sh1.cell(6,2).value)).grid(sticky=W)
        Radiobutton(self.logf,text = sh1.cell(7,2).value,bd = 3 ,font = ('',15),padx=5,pady=5,command=lambda: self.trv( sh1.cell(7,2).value)).grid(sticky=W)
        Radiobutton(self.logf,text = sh1.cell(8,2).value,bd = 3 ,font = ('',15),padx=5,pady=5,command=lambda: self.trv( sh1.cell(8,2).value)).grid(sticky=W)
        Radiobutton(self.logf,text = sh1.cell(9,2).value,bd = 3 ,font = ('',15),padx=5,pady=5,command=lambda: self.trv( sh1.cell(9,2).value)).grid(sticky=W)
        Radiobutton(self.logf,text = sh1.cell(10,2).value,bd = 3 ,font = ('',15),padx=5,pady=5,command=lambda: self.trv( sh1.cell(10,2).value)).grid(sticky=W)
        #Radiobutton(self.logf,text = sh1.cell(11,2).value,bd = 3 ,font = ('',15),padx=5,pady=5,command=lambda: self.trv( sh1.cell(11,2).value)).grid(sticky=W)
        
        self.logf.pack()
    def Car(self):
        self.head.pack_forget()
        sh1=wb["Car"]
        self.head = Label(self.master,text = 'Choose A Car',font = ('',30),pady = 10)
        self.head.pack()
        self.logf = Frame(self.master,padx =10,pady = 10)
        Radiobutton(self.logf,text = sh1.cell(2,2).value,bd = 3 ,font = ('',15),padx=5,pady=5,command=lambda :self.trv( sh1.cell(2,2).value)).grid()
        Radiobutton(self.logf,text = sh1.cell(3,2).value,bd = 3 ,font = ('',15),padx=5,pady=5,command=lambda :self.trv( sh1.cell(3,2).value)).grid()
        Radiobutton(self.logf,text = sh1.cell(4,2).value,bd = 3 ,font = ('',15),padx=5,pady=5,command=lambda: self.trv( sh1.cell(4,2).value)).grid()
        Radiobutton(self.logf,text = sh1.cell(5,2).value,bd = 3 ,font = ('',15),padx=5,pady=5,command=lambda: self.trv( sh1.cell(5,2).value)).grid()
        Radiobutton(self.logf,text = sh1.cell(6,2).value,bd = 3 ,font = ('',15),padx=5,pady=5,command=lambda: self.trv( sh1.cell(6,2).value)).grid()
        Radiobutton(self.logf,text = sh1.cell(7,2).value,bd = 3 ,font = ('',15),padx=5,pady=5,command=lambda: self.trv( sh1.cell(7,2).value)).grid()
        Radiobutton(self.logf,text = sh1.cell(8,2).value,bd = 3 ,font = ('',15),padx=5,pady=5,command=lambda: self.trv( sh1.cell(8,2).value)).grid()
        Radiobutton(self.logf,text = sh1.cell(9,2).value,bd = 3 ,font = ('',15),padx=5,pady=5,command=lambda: self.trv( sh1.cell(9,2).value)).grid()
        Radiobutton(self.logf,text = sh1.cell(10,2).value,bd = 3 ,font = ('',15),padx=5,pady=5,command=lambda: self.trv( sh1.cell(10,2).value)).grid()
        Radiobutton(self.logf,text = sh1.cell(11,2).value,bd = 3 ,font = ('',15),padx=5,pady=5,command=lambda: self.trv( sh1.cell(11,2).value)).grid()
        #Radiobutton(self.logf,text = sh1.cell(12,2).value,bd = 3 ,font = ('',15),padx=5,pady=5,command=lambda: self.trv(sh1.cell(12,2).value)).grid()
        self.crof.pack_forget()
        self.logf.pack()
    def Cycle(self):
        self.head.pack_forget()
        sh1=wb["Cycle"]
        self.head = Label(self.master,text = 'Choose A Cycle',font = ('',30),pady = 10)
        self.head.pack()
        self.logf = Frame(self.master,padx =10,pady = 10)
        Radiobutton(self.logf,text = sh1.cell(2,2).value,bd = 3 ,font = ('',15),padx=5,pady=5,command=lambda :self.trv( sh1.cell(2,2).value)).grid()
        Radiobutton(self.logf,text = sh1.cell(3,2).value,bd = 3 ,font = ('',15),padx=5,pady=5,command=lambda :self.trv( sh1.cell(3,2).value)).grid()
        Radiobutton(self.logf,text = sh1.cell(4,2).value,bd = 3 ,font = ('',15),padx=5,pady=5,command=lambda: self.trv( sh1.cell(4,2).value)).grid()
        Radiobutton(self.logf,text = sh1.cell(5,2).value,bd = 3 ,font = ('',15),padx=5,pady=5,command=lambda: self.trv( sh1.cell(5,2).value)).grid()
        Radiobutton(self.logf,text = sh1.cell(6,2).value,bd = 3 ,font = ('',15),padx=5,pady=5,command=lambda: self.trv( sh1.cell(6,2).value)).grid()
        Radiobutton(self.logf,text = sh1.cell(7,2).value,bd = 3 ,font = ('',15),padx=5,pady=5,command=lambda: self.trv( sh1.cell(7,2).value)).grid()
        Radiobutton(self.logf,text = sh1.cell(8,2).value,bd = 3 ,font = ('',15),padx=5,pady=5,command=lambda: self.trv( sh1.cell(8,2).value)).grid()
        Radiobutton(self.logf,text = sh1.cell(9,2).value,bd = 3 ,font = ('',15),padx=5,pady=5,command=lambda: self.trv( sh1.cell(9,2).value)).grid()
        Radiobutton(self.logf,text = sh1.cell(10,2).value,bd = 3 ,font = ('',15),padx=5,pady=5,command=lambda: self.trv( sh1.cell(10,2).value)).grid()
        Radiobutton(self.logf,text = sh1.cell(11,2).value,bd = 3 ,font = ('',15),padx=5,pady=5,command=lambda: self.trv( sh1.cell(11,2).value)).grid()
        #Radiobutton(self.logf,text = sh1.cell(12,2).value,bd = 3 ,font = ('',15),padx=5,pady=5,command=lambda: self.trv(sh1.cell(12,2).value)).grid()
        self.crof.pack_forget()
        self.logf.pack()
    def Scooter(self):
        self.head.pack_forget()
        sh1=wb["Scooter"]
        self.head = Label(self.master,text = 'Choose A Scooter',font = ('',30),pady = 10)
        self.head.pack()
        self.logf = Frame(self.master,padx =10,pady = 10) 
        Checkbutton(self.logf,text = sh1.cell(2,2).value,bd = 3 ,font = ('',15),padx=5,pady=5,command=lambda :self.trv( sh1.cell(2,2).value)).grid()
        Radiobutton(self.logf,text = sh1.cell(3,2).value,bd = 3 ,font = ('',15),padx=5,pady=5,command=lambda :self.trv( sh1.cell(3,2).value)).grid()
        Radiobutton(self.logf,text = sh1.cell(4,2).value,bd = 3 ,font = ('',15),padx=5,pady=5,command=lambda: self.trv( sh1.cell(4,2).value)).grid()
        Radiobutton(self.logf,text = sh1.cell(5,2).value,bd = 3 ,font = ('',15),padx=5,pady=5,command=lambda: self.trv( sh1.cell(5,2).value)).grid()
        Radiobutton(self.logf,text = sh1.cell(6,2).value,bd = 3 ,font = ('',15),padx=5,pady=5,command=lambda: self.trv( sh1.cell(6,2).value)).grid()
        Radiobutton(self.logf,text = sh1.cell(7,2).value,bd = 3 ,font = ('',15),padx=5,pady=5,command=lambda: self.trv( sh1.cell(7,2).value)).grid()
        Radiobutton(self.logf,text = sh1.cell(8,2).value,bd = 3 ,font = ('',15),padx=5,pady=5,command=lambda: self.trv( sh1.cell(8,2).value)).grid()
        Radiobutton(self.logf,text = sh1.cell(9,2).value,bd = 3 ,font = ('',15),padx=5,pady=5,command=lambda: self.trv( sh1.cell(9,2).value)).grid()
        Radiobutton(self.logf,text = sh1.cell(10,2).value,bd = 3 ,font = ('',15),padx=5,pady=5,command=lambda: self.trv( sh1.cell(10,2).value)).grid()
        Radiobutton(self.logf,text = sh1.cell(11,2).value,bd = 3 ,font = ('',15),padx=5,pady=5,command=lambda: self.trv( sh1.cell(11,2).value)).grid()
        #Radiobutton(self.logf,text = sh1.cell(12,2).value,bd = 3 ,font = ('',15),padx=5,pady=5,command=lambda: self.trv(sh1.cell(12,2).value)).grid()
        self.crof.pack_forget()
        self.logf.pack()
        
    def trv (self,Vehicle):
        self.s1=wb["Bike"]
        self.s2=wb["Car"]
        self.s3=wb["Cycle"]
        self.s4=wb["Scooter"]
        self.head.pack_forget()
        self.logf.pack_forget()
        self.vehicle=Vehicle
        for a in range (1,12):
                    
            if self.vehicle ==self.s1.cell(a,2).value:
                Exit= ms.askyesno("Prompt",f"Serial Number:{self.s1.cell(a,1).value} \n Name:{self.s1.cell(a,2).value} \n Occupancy:{self.s1.cell(a,3).value}\n Please! Remember Serial Number")
                if Exit > 0:
                    # root.destroy()
                    afterselectingVehicle(root, self.s1.cell(a,2).value)
                    return
            if self.vehicle ==self.s2.cell(a,2).value:
                Exit= ms.askyesno("Prompt",f"Serial Number:{self.s2.cell(a,1).value} \n Name:{self.s2.cell(a,2).value} \n Occupancy:{self.s2.cell(a,3).value}\n Please! Remember Serial Number")
                if Exit > 0:
                    # root.destroy()
                    afterselectingVehicle(root, self.s2.cell(a,2).value)
                    return
            if self.vehicle ==self.s3.cell(a,2).value:
                Exit= ms.askyesno("Prompt",f"Serial Number:{self.s3.cell(a,1).value} \n Name:{self.s3.cell(a,2).value} \n Occupancy:{self.s3.cell(a,3).value}\n Please! Remember Serial Number")
                if Exit > 0:
                    # root.destroy()
                    afterselectingVehicle(root, self.s3.cell(a,2).value)
                    return
            if self.vehicle ==self.s4.cell(a,2).value:
                Exit= ms.askyesno("Prompt",f"Serial Number:{self.s4.cell(a,1).value} \n Name:{self.s4.cell(a,2).value} \n Occupancy:{self.s4.cell(a,3).value}\n Please! Remember Serial Number")
                if Exit > 0:
                    # root.destroy()
                    afterselectingVehicle(root, self.s4.cell(a,2).value)
                    return 

    def logout(self):
        Exit= ms.askyesno("Prompt!","Do you want to exit?")
        if Exit > 0:
            root.destroy()
            return


    def RAV(self):
        self.head.pack_forget()
        self.logf.pack_forget()
        self.logf = Frame(self.master,padx =10,pady = 10)
        
        Label(self.logf,text = 'Enter the Serial Number',font = ('',20),pady=5,padx=5).grid(sticky = W)
        Entry(self.logf,textvariable = self.number,bd = 5,font = ('',15)).grid()
        #Button(self.logf,text = ' Enter ',bd = 3 ,font = ('',15),padx=5,pady=5,command=lambda:self.ret(self.number.get())).grid()
        # self.btnExit = Button(self.logf,padx=18,bd=7,font=('arial',11,'bold'),width = 2,text='Exit', command=self.logout()).grid()
        
        Button(self.logf,text = ' Enter ',bd = 3 ,font = ('',15),padx=5,pady=5,command=lambda:self.checkSrNo(self.number.get())).grid()
        self.logf.pack()

    def checkSrNo(self, Number):
        self.logf.pack_forget()
        self.number = Number
        # wb=op.load_workbook("Vehicle.xlsx")
        self.s5 = wb["RentedVehicle"]
        for a in range (1,10):
            if self.number == str(self.s5.cell(a,6).value):
                ReturnVehicle(root, a)
                return
        else:
            #no Vehicles to return 
            n=ms.showerror("No Vehicles with this Serial Number Plz Enter Correct One ")
            if n== "ok":
                self.RAV()


 #====================================================================================================================================
######################hhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhh##################

class ReturnVehicle:
    def __init__(self, root, main_row):
        self.main_row = main_row
        self.root = root
        self.root.title("Vehicle Rental System")
        self.root.geometry(geometry) 
        self.root.configure(background='black')
        self.s5 = wb['RentedVehicle']

        # Firstname=StringVar()
        # Surname=StringVar()
        initial_reading1 = StringVar()
        final_reading1 = StringVar()
        
        Damage=StringVar()
        Damage.set("0")
        var1 = IntVar()

        def Cab_Tax():
            global Item1
            if var1.get() == 1:
                self.txtDamage.configure(state = NORMAL)
                Item1=float(5000)
                Damage.set(str(Item1))
            elif var1.get() == 0:
                self.txtDamage.configure(state=DISABLED)
                Damage.set("0")
                Item1=0

        def iExit():
            iExit= ms.askyesno("Prompt!","Do you want to exit?")
            if iExit > 0:
                root.destroy()
                return
        
        def Receiptt():
            final_read = int(final_reading1.get())
            initial_read = int(initial_reading1.get())
            Total_Reading = final_read - initial_read
            Total_bill = 2000+500+Total_Reading*10 + float(Damage.get())
            
            self.txtReceipt1.delete("1.0",END)
            self.txtReceipt2.delete("1.0",END)

            self.txtReceipt1.insert(END,"Receipt Ref:\n")
            self.txtReceipt2.insert(END, self.s5.cell(self.main_row,2).value + "\n")
            self.txtReceipt1.insert(END,'Date:\n')
            self.txtReceipt2.insert(END, self.s5.cell(self.main_row,1).value + "\n")
            self.txtReceipt1.insert(END,'Firstname:\n')
            self.txtReceipt2.insert(END, self.s5.cell(self.main_row,3).value + "\n")
            self.txtReceipt1.insert(END,'Surname:\n')
            self.txtReceipt2.insert(END, self.s5.cell(self.main_row,4).value + "\n")
            self.txtReceipt1.insert(END,'VehicleName: \n')
            self.txtReceipt2.insert(END, self.s5.cell(self.main_row,5).value + "\n")
            self.txtReceipt1.insert(END,'VehicleSerialNo: \n')
            self.txtReceipt2.insert(END, str(self.s5.cell(self.main_row,6).value) + "\n")
            self.txtReceipt1.insert(END,'Basic Charge: \n')
            self.txtReceipt2.insert(END, "2000 \n")
            self.txtReceipt1.insert(END,'Insaurance Charge: \n')
            self.txtReceipt2.insert(END, "500 \n")
            self.txtReceipt1.insert(END,'Total Reading: \n')
            self.txtReceipt2.insert(END, f"{Total_Reading} \n")

            self.txtReceipt1.insert(END,'Total Bill: \n')
            self.txtReceipt2.insert(END, f"{Total_bill} \n") # to be taken from calculation of bill

            #Now here will be code to remove this vehicle from Excel after payment
            self.s5.delete_rows(idx=self.main_row, amount=1)
            wb.save("Vehicle.xlsx")

    #================================================mainframe========================================================================

        MainFrame=Frame(self.root)
        MainFrame.pack(fill=BOTH,expand=True)
        
        Tops = Frame(MainFrame, bd=10, width=1350,relief=RIDGE)
        Tops.pack(side=TOP,fill=BOTH)

        self.lblTitle=Label(Tops,font=('arial',50,'bold'),text="\t   VEHICLE RENTAL SYSTEM ")
        self.lblTitle.grid()
    #================================================customerframedetail=============================================================
        CustomerDetailsFrame=LabelFrame(MainFrame, width=1350,height=500,bd=20, pady=5, relief=RIDGE)
        CustomerDetailsFrame.pack(side=BOTTOM,fill=BOTH,expand=True)

        FrameDetails=Frame(CustomerDetailsFrame, width=880,height=400,bd=10, relief=RIDGE)
        FrameDetails.pack(side=LEFT,fill=BOTH,expand=True)

        CustomerName=LabelFrame(FrameDetails, width=150,height=250,bd=10, font=('arial',12,'bold'),text="Customer Info", relief=RIDGE)
        CustomerName.grid(row=0,column=0)

        TravelFrame = LabelFrame(FrameDetails,bd=10, width=300,height=250, font=('arial',12,'bold'),text="Vehicle Detail", relief=RIDGE)
        TravelFrame.grid(row=0,column=1)

    #===============================================recipt======================================================================
        Receipt_BottonFrame=LabelFrame(CustomerDetailsFrame,bd=10, width=450,height=400, relief=RIDGE)
        Receipt_BottonFrame.pack(side=RIGHT,fill=BOTH,expand=True)

        ReceiptFrame=LabelFrame(Receipt_BottonFrame, width=450,height=300, font=('arial',12,'bold'),text="Receipt", relief=RIDGE)
        ReceiptFrame.grid(row=0,column=0)

        ButtonFrame=LabelFrame(Receipt_BottonFrame, width=350,height=100, relief=RIDGE)
        ButtonFrame.grid(row=1,column=0)

    #=========================================================CustomerName====================================================

        self.lblFirstname=Label(CustomerName,font=('arial',14,'bold'),text="Firstname",bd=7)
        self.lblFirstname.grid(row=0,column=0,sticky=W)
        self.txtFirstname=Label(CustomerName,font=('arial',14,'bold'),text=self.s5.cell(self.main_row,3).value,bd=7)
        self.txtFirstname.grid(row=0,column=1,sticky=W)


        self.lblSurname=Label(CustomerName,font=('arial',14,'bold'),text="Surname",bd=7)
        self.lblSurname.grid(row=1,column=0,sticky=W)
        self.txtSurname=Label(CustomerName,font=('arial',14,'bold'),text=self.s5.cell(self.main_row,4).value,bd=7)
        self.txtSurname.grid(row=1,column=1,sticky=W)

    #==================================================TravelFrame================================================================
        self.Initial_reading=Label(TravelFrame,font=('arial',14,'bold'),text="Initial reading",bd=7)
        self.Initial_reading.grid(row=0,column=0,sticky=W)
        self.inp_initial_reading=Entry(TravelFrame,font=('arial',14,'bold'),textvariable=initial_reading1,bd=7,insertwidth=2,justify=RIGHT)
        self.inp_initial_reading.grid(row=0,column=1)

        self.Final_reading=Label(TravelFrame,font=('arial',14,'bold'),text="Final Reading",bd=7)
        self.Final_reading.grid(row=1,column=0,sticky=W)
        self.inp_Final_reading=Entry(TravelFrame,font=('arial',14,'bold'),textvariable=final_reading1,bd=7,insertwidth=2,justify=RIGHT)
        self.inp_Final_reading.grid(row=1,column=1)

        self.lblBasic=Label(TravelFrame,font=('arial',14,'bold'),text="Basic Amount",bd=7)
        self.lblBasic.grid(row=3,column=0,sticky=W)
        self.txtBasic=Label(TravelFrame,font=('arial',14,'bold'),text='2000',bd=7)
        self.txtBasic.grid(row=3,column=1,sticky=W)

        self.lblInsaurance=Label(TravelFrame,font=('arial',14,'bold'),text="Insaurance Amount",bd=7)
        self.lblInsaurance.grid(row=4,column=0,sticky=W)
        self.txtInsaurance=Label(TravelFrame,font=('arial',14,'bold'),text='500',bd=7)
        self.txtInsaurance.grid(row=4,column=1,sticky=W)

        self.chkDamage=Checkbutton(TravelFrame,text="Check Damage",variable = var1, onvalue=1, offvalue=0,font=('arial',16,'bold'),command=Cab_Tax).grid(row=5, column=0, sticky=W)
        self.txtDamage=Label(TravelFrame,font=('arial',14,'bold'),textvariable=Damage,bd=6,width=18,bg="white",state= DISABLED,justify=RIGHT,relief=SUNKEN)
        self.txtDamage.grid(row=5,column=1)

        #=======Calculations of Bill======#
        #basic Amount = 2000
        #insurance =500
        #reading*10
        # Damage = 0
        # Total Bill = Basic_Amount + Insaurance + reading*10 + Damage
        


    #=======================================Receipt====================================================================================

        self.txtReceipt1 = Text(ReceiptFrame,width = 25, height = 21,font=('arial',10,'bold'),borderwidth=0)
        self.txtReceipt1.grid(row=0,column=0,columnspan=2)
        self.txtReceipt2 = Text(ReceiptFrame,width = 25, height = 21,font=('arial',10,'bold'),borderwidth=0)
        self.txtReceipt2.grid(row=0,column=2,columnspan=2)

    #======================================Button========================================================================================
        
        self.btnReceipt = Button(ButtonFrame,padx=18,bd=7,font=('arial',11,'bold'),width = 10,text='Receipt and Make Payment',command=Receiptt).grid(row=0,column=1)
        
        self.btnExit = Button(ButtonFrame,padx=18,bd=7,font=('arial',11,'bold'),width = 2,text='Exit', command=iExit).grid(row=0,column=3)
        


############hhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhh##########
class afterselectingVehicle():
    def __init__(self,root, Vehicle):
        self.root = root
        self.Vehicle=Vehicle
        self.root.title("Vehicle Rental System")
        self.root.geometry(geometry) 
        self.root.configure(background='black')

        self.s1=wb["Bike"]
        self.s2=wb["Car"]
        self.s3=wb["Cycle"]
        self.s4=wb["Scooter"]
        self.s5=wb["RentedVehicle"]


        for a in range (2,12):
            if self.s1.cell(a,2).value == self.Vehicle:
                self.serialNum  = self.s1.cell(a, 1).value
            if self.s2.cell(a,2).value == self.Vehicle:
                self.serialNum  = self.s2.cell(a, 1).value
            if self.s3.cell(a,2).value == self.Vehicle:
                self.serialNum  = self.s3.cell(a, 1).value
            if self.s4.cell(a,2).value == self.Vehicle:
                self.serialNum  = self.s4.cell(a, 1).value
            


        DateofOrder=StringVar()
        DateofOrder.set(time.strftime(" %d / %m / %Y "))
        Receipt_Ref=StringVar()

        reset_counter=0
        Firstname=StringVar()
        Surname=StringVar()
        Address=StringVar()
        Postcode=StringVar()
        Mobile=StringVar()
        Telephone=StringVar()
        Email=StringVar()

        def iExit():
            iExit= ms.askyesno("Prompt!","Do you want to exit?")
            if iExit > 0:
                root.destroy()
                return

        def Reset():

            Firstname.set("")
            Surname.set("")
            Address.set("")
            Postcode.set("")
            Mobile.set("")
            Telephone.set("")
            Email.set("")

            self.txtReceipt1.delete("1.0",END)
            self.txtReceipt2.delete("1.0",END)
            
            self.reset_counter=1
 
        
        def Receiptt():
            if reset_counter == 0 and Firstname.get()!="" and Surname.get()!="" and Address.get()!="" and Postcode.get()!="" and Mobile.get()!="" and Telephone.get()!="" and Email.get()!="":
                self.txtReceipt1.delete("1.0",END)
                self.txtReceipt2.delete("1.0",END)
                x=random.randint(10853,500831)
                randomRef = str(x)
                Receipt_Ref.set(randomRef)

                self.txtReceipt1.insert(END,"Ref:\n")
                self.txtReceipt2.insert(END, Receipt_Ref.get() + "\n")
                self.txtReceipt1.insert(END,'Date:\n')
                self.txtReceipt2.insert(END, DateofOrder.get() + "\n")
                self.txtReceipt1.insert(END,'Receipt No:\n')
                self.txtReceipt2.insert(END, 'TR ' + Receipt_Ref.get() + " BW\n")
                self.txtReceipt1.insert(END,'Firstname:\n')
                self.txtReceipt2.insert(END, Firstname.get() + "\n")
                self.txtReceipt1.insert(END,'Surname:\n')
                self.txtReceipt2.insert(END, Surname.get() + "\n")
                self.txtReceipt1.insert(END,'Address:\n')
                self.txtReceipt2.insert(END, Address.get() + "\n")
                self.txtReceipt1.insert(END,'Postal Code:\n')
                self.txtReceipt2.insert(END, Postcode.get() + "\n")
                self.txtReceipt1.insert(END,'Telephone:\n')
                self.txtReceipt2.insert(END, Telephone.get() + "\n")
                self.txtReceipt1.insert(END,'Mobile:\n')
                self.txtReceipt2.insert(END, Mobile.get() + "\n")
                self.txtReceipt1.insert(END,'Email:\n')
                self.txtReceipt2.insert(END, Email.get() + "\n")
                self.txtReceipt1.insert(END, 'Vehicle Name: \n')
                self.txtReceipt2.insert(END, f'{self.Vehicle}\n')
                self.txtReceipt1.insert(END, 'Serial Number: \n')
                self.txtReceipt2.insert(END, f'{self.serialNum}\n')
                
                s5=wb["RentedVehicle"]
                maxrows = s5.max_row
                s5.cell(maxrows+1, 1).value = DateofOrder.get()
                s5.cell(maxrows+1, 2).value = Receipt_Ref.get()
                s5.cell(maxrows+1, 3).value = Firstname.get()
                s5.cell(maxrows+1, 4).value = Surname.get()
                s5.cell(maxrows+1, 5).value = self.Vehicle
                s5.cell(maxrows+1, 6).value = self.serialNum
                wb.save("Vehicle.xlsx")

                
            else:
                self.txtReceipt1.delete("1.0",END)
                self.txtReceipt2.delete("1.0",END)
                self.txtReceipt1.insert(END,"\nNo Input")



    #================================================mainframe========================================================================

        MainFrame=Frame(self.root)
        MainFrame.pack(fill=BOTH,expand=True)
        
        Tops = Frame(MainFrame, bd=10, width=1350,relief=RIDGE)
        Tops.pack(side=TOP,fill=BOTH)

        self.lblTitle=Label(Tops,font=('arial',50,'bold'),text="\t   VEHICLE RENTAL SYSTEM ")
        self.lblTitle.grid()


    #================================================customerframedetail=============================================================
        CustomerDetailsFrame=LabelFrame(MainFrame, width=1350,height=500,bd=20, pady=5, relief=RIDGE)
        CustomerDetailsFrame.pack(side=BOTTOM,fill=BOTH,expand=True)

        FrameDetails=Frame(CustomerDetailsFrame, width=880,height=400,bd=10, relief=RIDGE)
        FrameDetails.pack(side=LEFT,fill=BOTH,expand=True)

        CustomerName=LabelFrame(FrameDetails, width=150,height=250,bd=10, font=('arial',12,'bold'),text="Customer Info", relief=RIDGE)
        CustomerName.grid(row=0,column=0)

        TravelFrame = LabelFrame(FrameDetails,bd=10, width=300,height=250, font=('arial',12,'bold'),text="Vehicle Detail", relief=RIDGE)
        TravelFrame.grid(row=0,column=1)

        Book_Frame=LabelFrame(FrameDetails,width=300,height=150,relief=FLAT)
        Book_Frame.grid(row=1,column=0)

        CostFrame = LabelFrame(FrameDetails,width=150,height=150,bd=5,relief=FLAT)
        CostFrame.grid(row=1,column=1)

    #===============================================recipt======================================================================
        Receipt_BottonFrame=LabelFrame(CustomerDetailsFrame,bd=10, width=450,height=400, relief=RIDGE)
        Receipt_BottonFrame.pack(side=RIGHT,fill=BOTH,expand=True)

        ReceiptFrame=LabelFrame(Receipt_BottonFrame, width=350,height=300, font=('arial',12,'bold'),text="Receipt", relief=RIDGE)
        ReceiptFrame.grid(row=0,column=0)

        ButtonFrame=LabelFrame(Receipt_BottonFrame, width=350,height=100, relief=RIDGE)
        ButtonFrame.grid(row=1,column=0)

    #=========================================================CustomerName====================================================

        self.lblFirstname=Label(CustomerName,font=('arial',14,'bold'),text="Firstname",bd=7)
        self.lblFirstname.grid(row=0,column=0,sticky=W)
        self.txtFirstname=Entry(CustomerName,font=('arial',14,'bold'),textvariable=Firstname,bd=7,insertwidth=2,justify=RIGHT)
        self.txtFirstname.grid(row=0,column=1)


        self.lblSurname=Label(CustomerName,font=('arial',14,'bold'),text="Surname",bd=7)
        self.lblSurname.grid(row=1,column=0,sticky=W)
        self.txtSurname=Entry(CustomerName,font=('arial',14,'bold'),textvariable=Surname,bd=7,insertwidth=2,justify=RIGHT)
        self.txtSurname.grid(row=1,column=1,sticky=W)


        self.lblAddress=Label(CustomerName,font=('arial',14,'bold'),text="Address",bd=7)
        self.lblAddress.grid(row=2,column=0,sticky=W)
        self.txtAddress=Entry(CustomerName,font=('arial',14,'bold'),textvariable=Address,bd=7,insertwidth=2,justify=RIGHT)
        self.txtAddress.grid(row=2,column=1)


        self.lblPostcode=Label(CustomerName,font=('arial',14,'bold'),text="Postcode",bd=7)
        self.lblPostcode.grid(row=3,column=0,sticky=W)
        self.txtPostcode=Entry(CustomerName,font=('arial',14,'bold'),textvariable=Postcode,bd=7,insertwidth=2,justify=RIGHT)
        self.txtPostcode.grid(row=3,column=1)


        self.lblTelephone=Label(CustomerName,font=('arial',14,'bold'),text="Telephone",bd=7)
        self.lblTelephone.grid(row=4,column=0,sticky=W)
        self.txtTelephone=Entry(CustomerName,font=('arial',14,'bold'),textvariable=Telephone,bd=7,insertwidth=2,justify=RIGHT)
        self.txtTelephone.grid(row=4,column=1)

        self.lblMobile=Label(CustomerName,font=('arial',14,'bold'),text="Mobile",bd=7)
        self.lblMobile.grid(row=5,column=0,sticky=W)
        self.txtMobile=Entry(CustomerName,font=('arial',14,'bold'),textvariable=Mobile,bd=7,insertwidth=2,justify=RIGHT)
        self.txtMobile.grid(row=5,column=1)

        self.lblEmail=Label(CustomerName,font=('arial',14,'bold'),text="Email",bd=7)
        self.lblEmail.grid(row=6,column=0,sticky=W)
        self.txtEmail=Entry(CustomerName,font=('arial',14,'bold'),textvariable=Email,bd=7,insertwidth=2,justify=RIGHT)
        self.txtEmail.grid(row=6,column=1)

    #======================================Booking Detail============================================================================##
        self.VehicleName=Label(TravelFrame,font=('arial',14,'bold'),text="Vehicle Name",bd=7)
        self.VehicleName.grid(row=0,column=0,sticky=W)
        self.inpVehicleName=Label(TravelFrame,font=('arial',14,'bold'),text=self.Vehicle,bd=7)
        self.inpVehicleName.grid(row=0,column=1,sticky=W)

        self.SerialNumber=Label(TravelFrame,font=('arial',14,'bold'),text="SerialNumber",bd=7)
        self.SerialNumber.grid(row=1,column=0,sticky=W)
        self.inpSerialNumber=Label(TravelFrame,font=('arial',14,'bold'),text=self.serialNum,bd=7)
        self.inpSerialNumber.grid(row=1,column=1,sticky=W)




    #=======================================Receipt====================================================================================

        self.txtReceipt1 = Text(ReceiptFrame,width = 22, height = 21,font=('arial',10,'bold'),borderwidth=0)
        self.txtReceipt1.grid(row=0,column=0,columnspan=2)
        self.txtReceipt2 = Text(ReceiptFrame,width = 22, height = 21,font=('arial',10,'bold'),borderwidth=0)
        self.txtReceipt2.grid(row=0,column=2,columnspan=2)

    #======================================Button========================================================================================
        
        self.btnReceipt = Button(ButtonFrame,padx=18,bd=7,font=('arial',11,'bold'),width = 2,text='Receipt',command=Receiptt).grid(row=0,column=1)
        self.btnReset = Button(ButtonFrame,padx=18,bd=7,font=('arial',11,'bold'),width = 2,text='Reset',command=Reset).grid(row=0,column=2)
        self.btnExit = Button(ButtonFrame,padx=18,bd=7,font=('arial',11,'bold'),width = 2,text='Exit', command=iExit).grid(row=0,column=3)
        

##########################################hhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhh###### 
      
if __name__=='__main__':
    root = Tk()

    #=========================================== Getting Screen Width ==================================================================
    w = root.winfo_screenwidth()
    h = root.winfo_screenheight()
    geometry="%dx%d+%d+%d"%(w,h,0,0)
    
    root.geometry("500x300+320+200")
    root.title('Login Form')
    root.wm_iconbitmap("VRS.ico")
    application = user(root)
    root.mainloop()
    
